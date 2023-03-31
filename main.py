import streamlit as st
import pandas as pd
import openpyxl

def QA(num):
    # 定义问题和选项
    question_1 = "Comparing the two full faces (Left and Right), which one looks more realistic?"
    options_1 = ["The Left one looks more realistic", "The Right one looks more realistic"]
    question_2 = "Comparing the lips of two faces, which one is more in sync with audio?"
    options_2 = ["The Left one is more in sync with audio", "The Right one is more in sync with audio"]

    # 显示问题并获取用户的答案
    answer_1 = st.radio(label=question_1, options=options_1, key=fr"button{num}.1")
    answer_2 = st.radio(label=question_2, options=options_2, key=fr"button{num}.2")
    ans1 = get_ans(answer_1)
    ans2 = get_ans(answer_2)

    output(num, ans1, ans2)

def get_ans(answer_str):
    if "Left" in answer_str:
        return "Left"
    elif "Right" in answer_str:
        return "Right"

def output(num, ans1, ans2):
    # 打开Excel文件
    book = openpyxl.load_workbook('test.xlsx')
    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    writer.book = book

    # 写入数据
    sheet = book.active
    sheet.cell(row=num+1, column=1, value=ans1)
    sheet.cell(row=num+1, column=2, value=ans2)

    # 关闭writer
    writer.save()
    writer.close()
    # 重新打开writer，并与Excel文件绑定
    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    writer.book = book

    # 关闭Excel文件
    book.close()

def play_video(num):
    st.subheader(fr"video{num}")
    st.video(fr'video_syn\{num}.mp4')
    st.write("Please answer the following questions, after you watch the video. ")

def instrunction():
    st.subheader("Instructions: ")
    text1 = 'Please watch the four short videos (duration 4~7s) of two animated talking heads. \
            You need to choose the talking head (the left or the right) that moves more naturally in terms of the full face and the lips. '
    text2 = '**Reminder 1**: Please **turn on the sound on your computer** while you are watching the videos. '
    text3 = '**Reminder 2**: Some of the videos (one or two) are qualification testing videos.\
             **Your task might get rejected if you make the choices randomly**.'
    st.write(text1)
    st.write(text2)
    st.write(text3)

instrunction()
# 创建DataFrame并保存为Excel文件
df = pd.DataFrame({'face': [''], 'lip': ['']})
df.to_excel('test.xlsx', index=False)

for i in range(1,33):
    play_video(i)
    QA(i)




