import streamlit as st
import pandas as pd
import openpyxl
import io

def QA(num):
    # 定义问题和选项
    question_1 = "Comparing the two full faces (Left and Right), which one looks more realistic?"
    options_1 = ["The Left one looks more realistic", "The Right one looks more realistic"]
    question_2 = "Comparing the lips of two faces, which one is more in sync with audio?"
    options_2 = ["The Left one is more in sync with audio", "The Right one is more in sync with audio"]

    # 显示问题并获取用户的答案
    answer_1 = st.radio(label=question_1, options=options_1, key=fr"button{num}.1")
    answer_2 = st.radio(label=question_2, options=options_2, key=fr"button{num}.2")

    ans1 = get_ans(answer_1)
    ans2 = get_ans(answer_2)

    # 在excel保存结果
    output(num, ans1, ans2)

def get_ans(answer_str):
    if "Left" in answer_str:
        return "1"
    elif "Right" in answer_str:
        return "0"

def output(num, ans1, ans2):
    # 打开Excel文件
    book = openpyxl.load_workbook('data_left1_right0.xlsx')
    writer = pd.ExcelWriter("data_left1_right0.xlsx", engine='openpyxl')
    writer.book = book

    # 写入数据
    sheet = book.active
    sheet.cell(row=num+1, column=1, value=ans1)
    sheet.cell(row=num+1, column=2, value=ans2)

    # 关闭writer
    writer.save()
    writer.close()

    # 关闭Excel文件
    book.close()

def play_video(num):
    st.subheader(fr"video{num}")
    st.video(fr'{num}.mp4')
    st.write("Please answer the following questions, after you watch the video. ")

def instrunction():
    st.subheader("Instructions: ")
    text1 = 'Please watch the four short videos (duration 4~7s) of two animated talking heads. \
            You need to choose the talking head (the left or the right) that moves more naturally in terms of the full face and the lips. '
    text2 = '**Reminder 1**: Please **turn on the sound on your computer** while you are watching the videos. '
    text3 = '**Reminder 2**: Some of the videos (one or two) are qualification testing videos.\
             **Your task might get rejected if you make the choices randomly**.'
    st.write(text1)
    st.write(text2)
    st.write(text3)


# 注意事项
instrunction()

# 定义页面跳转函数，同时清空页面内容
def switch_page(page_num):
    st.session_state["page_num"] = page_num
    st.experimental_rerun()

# 通过 st.session_state 实现页面跳转
if "page_num" not in st.session_state:
    st.session_state["page_num"] = 1
    # 创建一个空的 DataFrame
    df = pd.DataFrame({'face': [''], 'lip': ['']})

    # 在 DataFrame 中添加 33 行空数据
    for i in range(33):
        df.loc[i] = ['', '']

    # 在第二行第一列和第二列中写入数字 1
    df.loc[1, 'face'] = "1"
    df.loc[1, 'lip'] = "1"
    df.loc[0:31, ['face', 'lip']] = "1"  # 在第2行到第33行中，第1列和第2列都写入数字 1

    # 将 DataFrame 写入 Excel 文件
    df.to_excel('data_left1_right0.xlsx', index=False)

num = st.session_state["page_num"]
# 显示页面内容
play_video(num)
QA(num)

# 显示上一页和下一页按钮
# 第2页到31页
if num > 1 and num < 32:
    col1, col2 = st.columns(2)
    if col2.button("Previous"):
        switch_page(st.session_state["page_num"] - 1)
    if col1.button("Next"):
        switch_page(st.session_state["page_num"] + 1)

# 第1页和第32页
if st.session_state["page_num"] == 32:
    col1, col2 = st.columns(2)
    if col1.button("Submit results"):
        # 读取Excel文件
        df = pd.read_excel("data_left1_right0.xlsx")
        # 显示文件内容
        st.write(df)
        # 生成Excel文件链接
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
            download2 = st.download_button(
                label="Download data as Excel",
                data=buffer,
                file_name='data_left1_right0.xlsx',
                mime='application/vnd.ms-excel'
             )
            writer.close()
    if col2.button("Previous"):
        switch_page(st.session_state["page_num"] - 1)


if st.session_state["page_num"] == 1:
    if st.button("Next"):
        switch_page(st.session_state["page_num"] + 1)

