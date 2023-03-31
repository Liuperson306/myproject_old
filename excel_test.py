import pandas as pd
import openpyxl

def output(num, ans1, ans2):
    # 创建DataFrame并保存为Excel文件
    df = pd.DataFrame({'face': [''], 'lip': ['']})
    df.to_excel('test.xlsx', index=False)

    # 打开Excel文件
    book = openpyxl.load_workbook('test.xlsx')
    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    writer.book = book

    # 进行写操作
    for i in range(10):
        # 写入i和i-1
        sheet = book.active
        sheet.cell(row=i+2, column=1, value=ans1)
        sheet.cell(row=i+2, column=2, value=ans2)
        # 关闭writer
        writer.save()
        writer.close()
        # 重新打开writer，并与Excel文件绑定
        writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
        writer.book = book

    # 关闭Excel文件
    book.close()
